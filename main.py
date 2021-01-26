#!/Library/Frameworks/Python.framework/Versions/3.6/bin/python3.6
import os
import sys
import boto3
import openpyxl
#from PIL import Image, ImageChops



def trim(im):
	img = Image.open(im)
	img = img.convert('RGB')
	bg = Image.new(img.mode, img.size, img.getpixel((0,0)))
	diff = ImageChops.difference(img, bg)
	diff = ImageChops.add(diff, diff, 2.0, -100)
	bbox = diff.getbbox()
	if bbox:
		return img.crop(bbox).save(im)


def aws_polly(text, data_type):
	client = boto3.client('polly')
	if data_type == "json":
		response = client.synthesize_speech(
			LanguageCode='en-IN',
			OutputFormat='json',
			Text=text,
			SpeechMarkTypes=['word'],
			VoiceId='Raveena'
		)
	elif data_type == "mp3":
		response = client.synthesize_speech(
			LanguageCode='en-IN',
			OutputFormat='mp3',
			Text=text,
			VoiceId='Raveena'
		)

	return(response['AudioStream'])


def polly_audio(text):
	audio_data = aws_polly(text, "mp3")
	return(audio_data)


def polly_json(text):
	json_data = aws_polly(text, "json")
	return(json_data)


def underline_html(text):
	html = ""
	first = True
	for i in text:
		if i != '_':
			html += i
		else:
			if first:
				html += '<u>'
				first = False
			else:
				html += '</u>'
				first = True
	return html


def create_images(text, image):
	no_of_lines = len(text.strip().split('\n'))
	text = underline_html(text)
	text = text.strip()
	words = text.split(' ')
	# print(words)
	images = []
	for i in range(len(words)):
		text_html = ''
		if words[i][0] in "-:#\n":
				continue

		for j in range(len(words)):
			if words[j][0] == "#":
				text_html += "&nbsp;"*len(words[j])
				continue

			if words[j] == '\n':
				text_html += '<br>'
				continue

			if j == i:
				if words[j][-1] in ":":
					text_html += '<span style="color:red;">' + words[j][0:-1] + '</span> ' + words[j][-1]
				else:
					text_html += '<span style="color:red;">' + words[j] + '</span> '
			else:
				text_html += words[j] + ' '

		img = '<div style="height: 400px; display: flex; justify-content: center; align-items: flex-end; padding: 10px"><img src="{}" style="height: 350px;"></img></div>'.format(image)
		br = '<br>'

		if type(image) == type(''):
			# trim(image)
			text = '<h1 style="font-size: {0}rem; margin: 0px"><p style="margin: 0.5em; text-align:center">{1}</p></h1>'.format(7-no_of_lines, text_html)
			html = '<!DOCTYPE html><html><body style="margin:0px"><div id="vid_area" style="height: 720px; width: 1280px;">{0}<div style="height: 360px; display: flex; justify-content: center; align-items: flex-start;">{1}</div></div></body></html>'.format(img, text)
		else:
			text = '<div style="height: 720px; display: flex; justify-content: center; align-items: center;"><p style="margin: 0.5em; text-align:center">{}</p></div>'.format(text_html)
			html = '<!DOCTYPE html><html><body style="margin:0px"><div id="vid_area" style="height: 720px; width: 1280px;"><h1 style="font-size: 7rem">{0}</h1></div></body></html>'.format(text)

		with open("tmp.html", "w") as f:
			f.write(html)

		cmd = "node pup.js file://{0}/tmp.html images/{1}.jpg".format(os.getcwd(), str(i))
		os.system(cmd)

		images.append("images/{}.jpg".format(i))
	
	return(images)


def concatenate_videos(videos, output_name):
	with open('con.in', 'w') as f:
		for video in videos:
			f.write("file '{}'\n".format(video))
	os.system("ffmpeg -y -f concat -safe 0 -i con.in -strict -2 -video_track_timescale 90000 -max_muxing_queue_size 2048 -tune animation -crf 6 {}".format(output_name))


def create_para_vid(speed, i, time_data, images, audio, output_name):
	end_times = []
	time_data_l = 0
	for l in time_data.iter_lines():
		time = int(l.decode().split(",")[0][8:])/(1000) * (1/speed)
		time_data_l +=1
		end_times.append(time)

	end_times = end_times[1:]

	with open("ffmp.in", "w") as f:
		prev_time = 0.0
		for j in range(len(images)-1):
			duration = float(end_times[j]) - prev_time
			f.write("file {0}' \n".format(images[j]))
			f.write("duration {} \n".format(duration))
			prev_time = float(end_times[j])

		f.write("file '{0}' \n".format(images[-1]))
		f.write("duration {} \n".format(prev_time+0.5))
	os.system("ffmpeg -y -i {0} -f concat -i ffmp.in -strict -2 -vsync vfr -pix_fmt yuv420p -vf fps=24 -video_track_timescale 90000 -max_muxing_queue_size 2048 -tune animation -crf 6 {1}nf.mp4".format(audio, output_name))
	with open('blank.in', 'w') as f:
		f.write('\n'.join(["file '{0}'".format(images[-1]), "duration {}".format("1.3")]))
	os.system("ffmpeg -y -i {0} -f concat -i blank.in -strict -2 -vsync vfr -pix_fmt yuv420p -vf fps=24 -video_track_timescale 90000 -max_muxing_queue_size 2048 -tune animation -crf 6 fill{1}.mp4".format("blank_long.mp3", str(i)))
	output_name_mp4 = output_name + '.mp4'
	concatenate_videos(['{0}nf.mp4'.format(output_name), 'fill{}.mp4'.format(i)], output_name_mp4)


def create_intro_video(sheet):
	text = (sheet.cell(row=2, column=2).value).split('\n')
	audio_data = polly_audio('. '.join(text))
	with open('audio_uf.mp3', 'wb') as f:
			f.write(audio_data.read())
	os.system("ffmpeg -y -i {} -ar 48000 {}".format("audio_uf.mp3", "intro_audio.mp3"))
	with open('intro.in', 'w') as f:
		f.write('\n'.join(["file \'images/{0}.jpg\'".format("intro"), "duration {}".format(5)]))
	headers = ''
	for line in text:
		headers += '<h1 style="font-size:2.5rem">{}</h1>'.format(line)
	html = '<!DOCTYPE html><html><body><div id="vid_area" style="height: 720px; width:1280px; display: flex; flex-direction: column; justify-content: center; align-items: center;">{}</div></body></html>'.format(headers)
	with open("intro.html", "w") as f:
			f.write(html)
	os.system("node pup.js file://{0}/intro.html images/{1}.jpg".format(os.getcwd(), "intro"))
	os.system("ffmpeg -y -i {0} -f concat -i intro.in -strict -2 -vsync vfr -pix_fmt yuv420p -vf fps=24 -video_track_timescale 90000 -max_muxing_queue_size 2048 -tune animation -crf 6 {1}.mp4".format("intro_audio.mp3", "intronf"))
	with open('blank.in', 'w') as f:
		f.write('\n'.join(["file \'images/{0}.jpg\'".format("intro"), "duration {}".format("2.5")]))
	os.system("ffmpeg -y -i {0} -f concat -i blank.in -strict -2 -vsync vfr -pix_fmt yuv420p -vf fps=24 -video_track_timescale 90000 -max_muxing_queue_size 2048 -tune animation -crf 6 fill{1}.mp4".format("blank.mp3", "intro"))
	concatenate_videos(["intronf.mp4", "fillintro.mp4"], "intro.mp4")
	return("intro.mp4")


def read_excel(path, sheet):
	wb = openpyxl.load_workbook(path)
	return(wb[sheet])


def create_vids_from_excel(file_dir, sheet):
	os.system('mkdir images')
	sheet = read_excel(file_dir, sheet)
	create_intro_video(sheet)
	normal_videos = []
	slow_videos = []
	split_videos = []
	start = 3
	end = sheet.max_row
	for i in range(start, end):
		if not sheet.cell(row=i, column=1).value:
			continue
		para = sheet.cell(row=i, column=2).value
		polly_para = para.replace('*', '.').replace('#', '').replace('_', '')
		json_data = polly_json(polly_para)
		json_data_slow = polly_json(polly_para)
		audio_data = polly_audio(polly_para)
		with open('audio_uf.mp3', 'wb') as f:
			f.write(audio_data.read())
		os.system("ffmpeg -y -i {} -ar 48000 {}".format("audio_uf.mp3", "audio.mp3"))
		os.system("sox audio.mp3 audio_slow.mp3 tempo 0.75")
		split_para = '. '.join(polly_para.split())
		json_data_split = polly_json(split_para)
		audio_data_split = polly_audio(split_para)
		with open('audio_uf.mp3', 'wb') as f:
			f.write(audio_data_split.read())
		os.system("ffmpeg -y -i {} -ar 48000 {}".format("audio_uf.mp3", "audio_split.mp3"))

		images_text = ' '.join(para.replace('*', '').replace('\n', ' \n ').split(' '))
		print(images_text)
		images = create_images(images_text, sheet.cell(row=i, column=3).value)

		create_para_vid(1, i-start-1, json_data, images, 'audio.mp3', "vid{}".format(str(i-start+1)))
		create_para_vid(0.75, i-2, json_data_slow, images, 'audio_slow.mp3', "vid{}-slow".format(str(i-2)))
		create_para_vid(1, i-2, json_data_split, images, 'audio_split.mp3', "vid{}-split".format(str(i-2)))

		normal_videos.append("vid{}.mp4".format(str(i-start+1)))
		slow_videos.append("vid{}-slow.mp4".format(str(i-start+1)))
		split_videos.append("vid{}-split.mp4".format(str(i-start+1)))

	os.system("mkdir final_videos")

	concatenate_videos(["intro.mp4"]+normal_videos, "final_videos/final.mp4")
	concatenate_videos(["intro.mp4"]+slow_videos, "final_videos/final_slow.mp4")
	concatenate_videos(["intro.mp4"]+split_videos, "final_videos/final_split.mp4")

	os.system("rm *.mp4")
	os.system("rm a*.mp3")
	os.system("rm *.in")
	os.system("rm *_a*.mp3")
	os.system("rm *.html")
	os.system("rm -r images")

#read_excel("input.xlsx")
create_vids_from_excel(str(sys.argv[1]), str(sys.argv[2]))